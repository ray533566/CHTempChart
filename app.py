import io
import zipfile
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy
import os

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "Result_Chart_template.xlsx")

EXPECTED_COLUMNS = [
    "SN", "WO", "Issue", "Equipment", "CH",
    "Operational Current ", "Oper Temperature", "Oper Wavelength", "Oper status",
    "Max Current", "Max Temperature", "Max Wavelength", "Max status"
]

st.set_page_config(
    page_title="SN Info → Result Chart Converter",
    page_icon="📊",
    layout="centered"
)

st.title("📊 SN Info → Result Chart Converter")
st.markdown(
    "Upload your **SN_Info_list.xlsx** (or a ZIP of CSV files) "
    "to generate a ready-to-download **Result_Chart.xlsx** with all analysis sheets updated."
)

# ── helper: load a DataFrame from uploaded file ──────────────────────────────

def load_dataframe(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        df = pd.read_excel(uploaded_file, sheet_name=0, dtype=str)
    elif name.endswith(".csv"):
        df = pd.read_csv(uploaded_file, dtype=str)
    elif name.endswith(".zip"):
        with zipfile.ZipFile(uploaded_file) as z:
            csv_names = [n for n in z.namelist() if n.lower().endswith(".csv")]
            if not csv_names:
                st.error("No CSV files found inside the ZIP.")
                return None
            frames = []
            for csv_name in sorted(csv_names):
                with z.open(csv_name) as f:
                    frames.append(pd.read_csv(f, dtype=str))
            df = pd.concat(frames, ignore_index=True)
    else:
        st.error("Unsupported file type. Please upload .xlsx, .csv, or .zip (containing CSVs).")
        return None
    return df


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Strip whitespace from column names and map to expected headers."""
    df.columns = [str(c).strip() for c in df.columns]
    return df


def validate_columns(df: pd.DataFrame) -> bool:
    stripped = [c.strip() for c in df.columns]
    missing = [c for c in EXPECTED_COLUMNS if c.strip() not in stripped]
    if missing:
        st.warning(
            f"⚠️  Some expected columns were not found: **{', '.join(missing)}**\n\n"
            "The file will still be processed — missing columns will appear blank."
        )
    return True


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def build_result_chart(df: pd.DataFrame) -> bytes:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Raw Data"]

    # ── 1. Find header row and clear old data ───────────────────────────────
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=15):
        for cell in row:
            if str(cell.value).strip().lower() == "sn":
                header_row_idx = cell.row
                break
        if header_row_idx:
            break

    if header_row_idx is None:
        header_row_idx = 1

    # Delete all data rows below header (keep header)
    max_row = ws.max_row
    if max_row > header_row_idx:
        ws.delete_rows(header_row_idx + 1, max_row - header_row_idx)

    # ── 2. Read header mapping from template ────────────────────────────────
    header_cells = [ws.cell(header_row_idx, col) for col in range(1, 14)]
    template_headers = [str(c.value).strip() if c.value else "" for c in header_cells]

    # Map df columns → template column positions
    col_map = {}
    for i, th in enumerate(template_headers):
        for dc in df.columns:
            if dc.strip() == th:
                col_map[i] = dc
                break

    # ── 3. Copy style from header row to use as row template ────────────────
    style_row = header_row_idx  # borrow header style for data rows

    df = df.fillna("")

    for row_idx, (_, row) in enumerate(df.iterrows(), start=header_row_idx + 1):
        for col_idx in range(len(template_headers)):
            cell = ws.cell(row=row_idx, column=col_idx + 1)
            dc = col_map.get(col_idx)
            val = row[dc] if dc else ""
            # Try numeric conversion for numeric-looking values
            if val != "":
                try:
                    num = float(val)
                    cell.value = num
                    continue
                except (ValueError, TypeError):
                    pass
            cell.value = val if val != "" else None

    # ── 4. Save to bytes ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── UI ───────────────────────────────────────────────────────────────────────

uploaded = st.file_uploader(
    "Upload SN Info List",
    type=["xlsx", "xls", "csv", "zip"],
    help="Accepts .xlsx, .csv, or a .zip archive containing CSV files"
)

if uploaded:
    with st.spinner("Reading file…"):
        df = load_dataframe(uploaded)

    if df is not None:
        df = normalize_columns(df)
        validate_columns(df)

        st.success(f"✅ Loaded **{len(df):,}** rows × **{len(df.columns)}** columns")

        with st.expander("Preview first 10 rows", expanded=False):
            st.dataframe(df.head(10), use_container_width=True)

        col1, col2 = st.columns(2)
        col1.metric("Total Rows", f"{len(df):,}")
        col2.metric("Unique SNs", f"{df['SN'].nunique():,}" if "SN" in df.columns else "—")

        st.divider()
        st.subheader("Generate Result Chart")

        if st.button("🚀 Build Result_Chart.xlsx", type="primary", use_container_width=True):
            with st.spinner("Writing data into template and generating file…"):
                try:
                    result_bytes = build_result_chart(df)
                    st.success("✅ Result_Chart.xlsx is ready!")
                    st.download_button(
                        label="⬇️  Download Result_Chart.xlsx",
                        data=result_bytes,
                        file_name="Result_Chart.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                    st.info(
                        "💡 Open the downloaded file in Excel. "
                        "The **Temperature Distribution** and **CH Distribution** sheets "
                        "will auto-recalculate once Excel opens the file."
                    )
                except Exception as e:
                    st.error(f"❌ Error generating file: {e}")
                    raise

st.divider()
st.caption(
    "**How it works:** Your uploaded data replaces the *Raw Data* sheet in the template. "
    "The *Temperature Distribution* and *CH Distribution* sheets contain live Excel formulas "
    "that recalculate automatically when you open the file in Excel."
)
